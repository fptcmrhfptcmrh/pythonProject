import numpy as np
from openpyxl import load_workbook
from matplotlib import pyplot as plt
file_name="D:\pyhton_pr\지진발생확률 - 2학년 확통세특\국내지진목록_2014-10-01_2024-05-17.xlsx"
workbook=load_workbook(file_name,data_only=True)
ws=workbook['Sheet1']
strength=[]
graphx=[]
graphy=[]
for i in range(4,1084):
  #규모 / 위도 / 경도
  strength.append(ws['C'+str(i)].value)
  graphx.append(float(ws['F'+str(i)].value[:-2]))
  graphy.append(float(ws['G'+str(i)].value[:-2]))
plt.scatter(graphy,graphx)
plt.xlabel("longitude (N)")
plt.ylabel("latitude (E)")
plt.show()
s_strength=list(map(float,sorted(list(set(strength)))))
val=[]
for i in s_strength:
  val.append(strength.count(i))
x=np.arange(len(s_strength))
plt.bar(x,val)
plt.xticks(x,s_strength)
plt.show()
def check_ord(s):
  if s=="위치":
    cnt=0
    a,b=map(float,input("위도/경도 : ").split())
    for i in range(len(graphx)):
      #근처 위도/경도에 해당하는 점의 개수를 세아림
      if a-0.25<=graphx[i]<=a+0.25 and b-0.25<=graphy[i]<=b+0.25:cnt+=1
    return cnt
  elif s=="규모":
    cnt=0
    k=float(input("규모 : "))
    for i in strength:
      if k==i: cnt+=1
    return cnt
  elif s=="위치 & 규모":
    cnt=0
    a,b=map(float,input("위도/경도 : ").split())
    k=float(input("규모 : "))
    for i in range(len(graphx)):
      if a-0.25<=graphx[i]<=a+0.25 and b-0.25<=graphy[i]<=b+0.25 and strength[i]==k:cnt+=1
    return cnt
  
while True:
  order=list(input("조건부확률  :  A | B \n확률 : A").split("|"))
  print(order)
  if len(order)==1:
    print(check_ord(order[0])/1080*100,check_ord(order[0]))
  else:
    print(check_ord("위치 & 규모")/check_ord(order[1]))
  plt.show()